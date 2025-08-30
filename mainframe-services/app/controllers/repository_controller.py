import io
import os
from datetime import datetime
from io import BytesIO
from typing import List, Optional, Dict, Tuple

import openpyxl
import pandas as pd
from azure.storage.blob import BlobServiceClient
from bson import ObjectId
from fastapi import HTTPException
from git import Repo
from loguru import logger
from motor.motor_asyncio import AsyncIOMotorDatabase

from app.config.settings import settings
from app.controllers.project_controller import check_project_access
from app.models.repository import Repository
from app.schemas.common_schema import ErrorResponse
from app.schemas.repository_schema import GetFilesByRepositoryResponse, GetFileContentResponse
from app.utils.azure_blob import AzureBlobStorage

azure_blob_storage = AzureBlobStorage()


async def create_repository(db: AsyncIOMotorDatabase, repository: Repository) -> Repository:
    """Create a new repository."""
    repository_dict = repository.model_dump(by_alias=True)
    repository_dict["_id"] = ObjectId(repository_dict["_id"])
    repository_dict["project_id"] = ObjectId(repository_dict["project_id"])
    result = await db.repositories.insert_one(repository_dict)
    repository.id = result.inserted_id
    return repository


async def get_repository(db: AsyncIOMotorDatabase, repository_id: str) -> Optional[Repository]:
    """Get a repository by ID."""
    repository = await db.repositories.find_one({"_id": ObjectId(repository_id)})
    if repository:
        return Repository(**repository)
    return None


async def check_repository_project_access(
        db: AsyncIOMotorDatabase,
        repository_id: str,
        user_id: str,
        required_role: str = None
) -> Repository:
    """Check project permission for repository access."""
    repository = await get_repository(db, repository_id)
    if not repository:
        raise HTTPException(
            status_code=404,
            detail=ErrorResponse(
                message="Repository not found",
                error_code="REPOSITORY_NOT_FOUND"
            ).model_dump()
        )

    # Check project permission
    await check_project_access(
        db,
        str(repository.project_id),
        user_id,
        required_role
    )

    return repository


async def get_repositories(
        db: AsyncIOMotorDatabase,
        user_id: str,
        skip: int = 0,
        limit: int = 100
) -> tuple[List[Repository], int]:
    """Get a list of repositories for projects user has access to."""
    pipeline = [
        # Match user
        {"$match": {"_id": ObjectId(user_id)}},
        # Unwind permissions
        {"$unwind": "$permissions"},
        # Get project IDs user has access to
        {
            "$group": {
                "_id": None,
                "project_ids": {"$push": "$permissions.project_id"}
            }
        },
        # Lookup repositories for these projects
        {
            "$lookup": {
                "from": "repositories",
                "let": {"project_ids": "$project_ids"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {"$in": ["$project_id", "$$project_ids"]}
                        }
                    }
                ],
                "as": "repositories"
            }
        },
        # Project total count and paginated repositories
        {
            "$project": {
                "total": {"$size": "$repositories"},
                "repositories": {"$slice": ["$repositories", skip, limit]}
            }
        }
    ]

    result = await db.users.aggregate(pipeline).to_list(1)
    if not result:
        return [], 0

    repositories = [Repository(**repo) for repo in result[0]["repositories"]]
    total = result[0]["total"]
    return repositories, total


async def get_repository_by_project(
        db: AsyncIOMotorDatabase,
        project_id: str
) -> Optional[Repository]:
    """Get a repository by project ID."""
    try:
        repository = await db.repositories.find_one({"project_id": ObjectId(project_id)})
        if repository:
            return Repository(**repository)
        return None
    except Exception as e:
        logger.error(f"Error getting repository for project {project_id}: {str(e)}")
        return None


async def update_repository(
        db: AsyncIOMotorDatabase,
        repository_id: str,
        update_data: dict,
) -> Optional[dict]:
    """
    Update a repository with partial data.
    
    Args:
        db: Database connection
        repository_id: Repository ID to update
        update_data: Dictionary containing only the fields to update
        
    Returns:
        Updated fields if successful, None if repository not found
    """
    try:
        # Convert project_id to ObjectId if it exists in update_data
        if 'project_id' in update_data:
            update_data['project_id'] = ObjectId(update_data['project_id'])

        result = await db.repositories.update_one(
            {"_id": ObjectId(repository_id)},
            {"$set": update_data}
        )

        if result.modified_count:
            return update_data
        return None

    except Exception as e:
        logger.error(f"Error updating repository {repository_id}: {str(e)}")
        return None


async def delete_repository(db: AsyncIOMotorDatabase, repository_id: str) -> bool:
    """Delete a repository."""
    result = await db.repositories.delete_one({"_id": ObjectId(repository_id)})
    return result.deleted_count > 0


async def process_repository(
        db: AsyncIOMotorDatabase,
        git_repo: Repo,
        repository: Repository
) -> None:
    """Process and upload repository files to Azure Blob Storage."""
    blob_service_client = BlobServiceClient.from_connection_string(
        settings.AZURE_STORAGE_CONNECTION_STRING
    )

    async def parse_tree(tree, repo_id, path):
        path = os.path.join(path, tree.name)

        for child in tree.trees:
            await parse_tree(child, repo_id, path)

        for blob in tree.blobs:
            local_file_name = os.path.join(str(repo_id), path, blob.name)
            upload_file_path = os.path.join("tmp", local_file_name)

            blob_client = blob_service_client.get_blob_client(
                container="repositories",
                blob=local_file_name
            )

            with open(file=upload_file_path, mode="rb") as data:
                blob_client.upload_blob(data)

            logger.info(f"Uploaded {blob.name} of {repository.id} to Azure Blob Storage")

    logger.info(f"Start uploading files from {repository.id} to Azure Blob Storage")
    await parse_tree(git_repo.tree(), repository.id, "")

    await update_repository(db, repository.id, repository.model_dump())

    logger.info(f"Uploaded all files from {repository.id} to Azure Blob Storage")


def get_parsed_csv(repo_id: str) -> io.BytesIO:
    """Get parsed CSV file from Azure Blob Storage."""
    blob_service_client = BlobServiceClient.from_connection_string(
        settings.AZURE_STORAGE_CONNECTION_STRING
    )
    container_client = blob_service_client.get_container_client(
        settings.AZURE_STORAGE_CONTAINER_NAME
    )
    blob_client = container_client.get_blob_client(f"{repo_id}_parsed.csv")
    blob_data = blob_client.download_blob().readall()
    return io.BytesIO(blob_data)


async def count_total_assessments(db, repository_id):
    pipeline = [
        {"$match": {"repository_id": ObjectId(repository_id)}},
        {"$project": {"total_assessments": {"$size": "$result.assessment"}}},
        {"$group": {"_id": None, "total_assessments": {"$sum": "$total_assessments"}}}
    ]

    result = await db.assessments.aggregate(pipeline).to_list(length=None)  # Use async to_list()
    total_assessments = result[0]["total_assessments"] if result else 0

    logger.info(f"Total Assessments for repository {repository_id}: {total_assessments}")
    return total_assessments


async def get_files_by_repo_id(db, repository_id: str, page: int, limit: int):
    skip = (page - 1) * limit
    files = []
    collected = 0

    cursor = db.assessments.find({"repository_id": ObjectId(repository_id)}, {"result.assessment": 1})

    async for doc in cursor:
        assessments = doc.get("result", {}).get("assessment", [])
        if skip >= len(assessments):
            skip -= len(assessments)
            continue

        for assessment in assessments[skip:]:
            if collected >= limit:
                return files
            files.append(assessment)
            collected += 1

        skip = 0

    return files


async def get_files_by_repository(db, repository_id, page_number, page_limit):
    total_assessments = await count_total_assessments(db, repository_id)
    files = await get_files_by_repo_id(db, repository_id, page_number, page_limit)
    return GetFilesByRepositoryResponse(files=files, total=total_assessments, skip=(page_number - 1) * page_limit,
                                        limit=page_limit)


async def get_file_by_blob_path(db, repository_id, file_type, file_name):
    try:
        pipeline = [
            {"$match": {
                "repository_id": ObjectId(repository_id),
            }},
            {"$unwind": "$result.assessment"},
            {"$match": {
                "result.assessment.name": file_name,
                "result.assessment.label": file_type
            }},
            {"$project": {
                "_id": 0,
                "path": "$result.assessment.path"
            }}
        ]

        blob_path = await db.assessments.aggregate(pipeline).to_list(1)
        if not blob_path or type(blob_path) != list:
            raise FileNotFoundError
        path = blob_path[0].get("path")
        content = await azure_blob_storage.read_blob(path)
        return GetFileContentResponse(title=path, content=content)
    except Exception as e:
        logger.error(f"Error getting file content: {e}")
        raise HTTPException(
            status_code=404,
            detail=ErrorResponse(
                message="File not found",
                error_code="FILE_NOT_FOUND"
            ).model_dump()
        )


async def generate_source_stats_report(
        db: AsyncIOMotorDatabase,
        repository_id: str
) -> Tuple[BytesIO, str]:
    """
    Generate a source code statistics report for a given repository.
    
    Args:
        db (AsyncIOMotorDatabase): MongoDB database connection
        repository_id (str): The ID of the repository to generate the report for
        
    Returns:
        Tuple[BytesIO, str]: A tuple containing:
            - BytesIO: The Excel file in memory
            - str: The filename for the report
            
    Raises:
        HTTPException: If repository_id is invalid or no assessments are found
    """
    try:
        # Validate repository_id format
        try:
            ObjectId(repository_id)
        except Exception:
            raise HTTPException(
                status_code=400,
                detail="Invalid repository_id format. Must be a valid MongoDB ObjectId."
            )

        # Get the latest assessment for the repository
        assessment = await db.assessments.find_one(
            {"repository_id": ObjectId(repository_id)},
            sort=[("created_at", -1)]
        )

        if not assessment or not assessment.get("result", {}).get("assessment"):
            raise HTTPException(
                status_code=404,
                detail=f"No assessment data found for repository {repository_id}"
            )

        # Extract statistics from assessment result
        assessments = assessment["result"]["assessment"]
        stats: Dict[str, Dict[str, int]] = {}
        rows = [] 

        # Process file statistics
        for file_data in assessments:
            file_type = file_data.get("label")
            if not file_type:
                continue

            # Update stats for "Summary"
            if file_type not in stats:
                stats[file_type] = {"nof": 0, "loc": 0}

            stats[file_type]["nof"] += 1
            stats[file_type]["loc"] += file_data.get("code", 0) if isinstance(file_data.get("code"), int) else 0

            # Prepare row data
            row = [
                file_data.get("name", "Unknown"),
                file_type,
                file_data.get("code", 0) if isinstance(file_data.get("code"), int) else 0
            ]
            # Append row to the list
            rows.append(row)

        # Convert to DataFrame
        data = []
        for file_type, type_stats in stats.items():
            data.append({
                "File Type": file_type,
                "Number of Files": type_stats["nof"],
                "Lines of code": type_stats["loc"]
            })

        # Sort by Type
        df = pd.DataFrame(data)
        df = df.sort_values("File Type")

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Source Statistics summary', index=False)

            # Get workbook and worksheet objects
            workbook = writer.book
            worksheet_summary = writer.sheets['Source Statistics summary']

            # Add some formatting
            for cell in worksheet_summary[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Set column widths
            for column_cells in worksheet_summary.columns:
                max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
                adjusted_width = max_length + 2
                worksheet_summary.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            
            worksheet_details = workbook.create_sheet(title='Source Statistics Details')
            # Add header
            headers = ["No", "Name", "Type", "LOC"]
            worksheet_details.append(headers)

            # Format header
            for cell in worksheet_details[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            rows.sort(key=lambda x: x[1])
            for idx, row in enumerate(rows, start=1):
                worksheet_details.append([idx] + row)

            for column_cells in worksheet_details.columns:
                max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
                adjusted_width = max_length + 5  
                worksheet_details.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Set the pointer to the beginning of the BytesIO object
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"source_stats_report_{timestamp}.xlsx"

        return output, filename

    except HTTPException as http_exc:
        logger.error(f"HTTPException occurred: {http_exc.detail}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error occurred while generating source statistics report: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error generating source statistics report: {str(e)}"
        )
